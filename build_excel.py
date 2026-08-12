#!/usr/bin/env python3
"""
MarketPano - Excel uretici (genis format)
=========================================
Cikti: marketpano.xlsx

  Sheet 1-6    : Binance, OKX, Bybit, Bitget, Gate, Hyperliquid
                 A sutunu Varlik. Sonra her GUN icin 4 sutunluk blok:
                 Fiyat | Market Cap | Perp Hacim | Open Interest
                 Yeni gun SAGA eklenir. Tarih ve Sira sutunu yoktur.

  Sheet 7      : Total
                 A sutunu Varlik, sonraki sutunlar tarih.
                 Deger: o varligin 6 borsadaki TOPLAM perp hacmi.

  Sheet 8-57   : Her varlik icin ayri kontrat sheet'i (CMC ilk 50).
                 Satirlar borsalar alt alta (+ en altta BTCTURK).
                 Sutunlar 12 kontrat alani.

  Sonraki      : Analiz - Varlik | Borsa | 12 alan, filtreli tek tablo.
                 Filtreden varlik secince o varligin tum borsalari alt alta.

  Son          : Notlar

Girdiler (bu dosyanin klasorunde):
  arsiv/hacim/YYYY-MM-DD.json      (yoksa hacim*.json tek gun)
  arsiv/cmc_genis/YYYY-MM-DD.json  (yoksa arsiv/cmc, yoksa cmc_genis.json/cmc.json)
  kontrat_github.json + kontrat_local.json   (kontrat TEK FOTOGRAF)
  manuel-kaldirac.csv, manuel-funding.csv, manuel-btcturk.csv

Excel her calistirmada SIFIRDAN uretilir; dosyaya elle yazilanlar korunmaz.
Elle girilecek veriler manuel-*.csv dosyalarinda tutulur.
"""
import csv
import glob
import json
import os
import re

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

os.chdir(os.path.dirname(os.path.abspath(__file__)))

OUT = "marketpano.xlsx"
BORSA_SIRASI = ["Binance", "OKX", "Bybit", "Bitget", "Gate", "Hyperliquid", "MEXC"]
KENDI_BORSA = "BTCTURK"

# Hacim sheet'lerinde her gun icin tekrarlanan sutunlar
# En guncel gun icin 4 sutun, gecmis gunler icin sadece perp hacim
GUNCEL_METRIK = ["Fiyat ($)", "Market Cap ($)", "Open Interest ($)", "Perp Hacim ($)"]
ESKI_METRIK = ["Perp Hacim ($)"]
GUN_DESEN = re.compile(r"^\d{4}-\d{2}-\d{2}$")

KONTRAT_SUTUNLAR = ["Index Price", "Min Miktar", "Min Tutar ($)", "Fiyat Adimi",
                    "Digit", "Miktar Adimi", "Max Kaldirac", "Funding %",
                    "Funding Periyot (saat)", "Derinlik +-1% ($)", "Spread %",
                    "Index Kirilimi"]

BORSA_RENK = {
    "Binance": "F0B90B", "OKX": "2B3139", "Bybit": "FF7A00",
    "Bitget": "7B61FF", "Gate": "E5402B", "Hyperliquid": "50D2C2",
    "MEXC": "1972F5", KENDI_BORSA: "003AFF",
}
BORSA_DOLGU = {
    "Binance": "FFFBF0", "OKX": "F2F3F5", "Bybit": "FFF6EE",
    "Bitget": "F6F3FF", "Gate": "FFF1EF", "Hyperliquid": "EFFBF9",
    "MEXC": "EEF4FE", KENDI_BORSA: "E6EDFF",
}

FONT = Font(name="Arial", size=10)
FONT_B = Font(name="Arial", size=10, bold=True)
HEAD_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
HEAD_FILL = PatternFill("solid", fgColor="003AFF")
ALT_FILL = PatternFill("solid", fgColor="EEF3FB")
ALT_FONT = Font(name="Arial", size=9, bold=True, color="35496B")
THIN = Side(style="thin", color="D6DEEA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

_CARPAN_RE = re.compile(r"^(?:1000000|100000|10000|1000|100|10|1M|1B)([A-Z0-9]{2,})$")
_K_RE = re.compile(r"^k([A-Z]{2,})$")


# ---------------- yardimcilar ----------------

def load_json(path):
    try:
        with open(path, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None


def temel_sembol(sym):
    """1000PEPE -> PEPE, kPEPE -> PEPE. Carpan yoksa None."""
    if not sym:
        return None
    m = _CARPAN_RE.match(sym) or _K_RE.match(sym)
    return m.group(1) if m else None


def gun_bicim(gun):
    """2026-07-30 -> 30.07.2026"""
    try:
        y, a, g = gun.split("-")
        return f"{g}.{a}.{y}"
    except Exception:
        return gun


def hacim_gunleri():
    """{gun: {borsa: [satirlar]}}  - arsiv yoksa eldeki dosyalardan tek gun."""
    gunler = {}
    for p in sorted(glob.glob("arsiv/hacim/*.json")):
        ad = os.path.basename(p)[:-5]
        if not GUN_DESEN.match(ad):      # index.json gibi dosyalari atla
            continue
        d = load_json(p)
        if d:
            gunler[ad] = d.get("borsalar") or {}
    if gunler:
        return gunler
    birlesik, gun = {}, None
    for p in ("hacim_github.json", "hacim_local.json", "hacim.json"):
        d = load_json(p)
        if not d:
            continue
        gun = gun or (d.get("generated_at") or "")[:10]
        birlesik.update(d.get("borsalar") or {})
    if birlesik and gun:
        print(f"  ! arsiv/hacim bos -> eldeki dosyalar tek gun olarak alindi ({gun}).")
        return {gun: birlesik}
    return {}


def cmc_gunleri():
    """{gun: {sembol: (fiyat, mcap)}} - once genis liste (1000), sonra 150."""
    gunler = {}
    for klasor in ("arsiv/cmc_genis", "arsiv/cmc"):
        for p in sorted(glob.glob(f"{klasor}/*.json")):
            gun = os.path.basename(p)[:-5]
            if not GUN_DESEN.match(gun):
                continue
            d = load_json(p)
            if not d:
                continue
            hedef = gunler.setdefault(gun, {})
            for c in d.get("coins", []):
                hedef.setdefault(c["symbol"], (c.get("price_usd"), c.get("market_cap_usd")))
    if gunler:
        return gunler
    for aday in ("cmc_genis.json", "cmc.json"):
        d = load_json(aday)
        if d:
            gun = (d.get("generated_at") or "")[:10]
            gunler[gun] = {c["symbol"]: (c.get("price_usd"), c.get("market_cap_usd"))
                           for c in d.get("coins", [])}
            break
    return gunler


def cmc_bak(cmc, gun, sembol):
    """O gunun fiyat+mcap'i. Gun yoksa en yakin onceki gun. Carpanli sembolde temel token."""
    def bak(tablo):
        if sembol in tablo:
            return tablo[sembol]
        t = temel_sembol(sembol)
        if t and t in tablo:
            return tablo[t]
        return None
    if gun in cmc:
        r = bak(cmc[gun])
        if r:
            return r
    for g in sorted((x for x in cmc if x <= gun), reverse=True):
        r = bak(cmc[g])
        if r:
            return r
    for g in sorted(cmc):
        r = bak(cmc[g])
        if r:
            return r
    return (None, None)


def btcturk_spot_seti():
    """btcturk_spot.json -> BtcTurk spot'ta listeli varlik kumesi."""
    d = load_json("btcturk_spot.json")
    if not d:
        print("  ! btcturk_spot.json yok -> BtcTurk Spot sutunu bos kalacak.")
        return None
    return set(d.get("varliklar") or [])


def kontrat_snapshot():
    """Kontrat TEK FOTOGRAF: kontrat_github.json + kontrat_local.json.
    Doner: (varlik listesi, {borsa: cekildigi_gun})"""
    birlesik, tarihler = {}, {}
    for p in ("kontrat_github.json", "kontrat_local.json", "kontrat.json"):
        d = load_json(p)
        if not d:
            continue
        gun = (d.get("generated_at") or "")[:10]
        for a in d.get("assets", []):
            rec = birlesik.setdefault(a["symbol"], {
                "symbol": a["symbol"], "rank": a.get("rank"), "exchanges": {}})
            if rec.get("rank") is None:
                rec["rank"] = a.get("rank")
            for borsa, v in (a.get("exchanges") or {}).items():
                rec["exchanges"][borsa] = v
                tarihler[borsa] = gun
    return sorted(birlesik.values(), key=lambda a: a.get("rank") or 9999), tarihler


def csv_oku(dosya, anahtar):
    """CSV -> {anahtar_degeri: satir_dict}"""
    out = {}
    try:
        with open(dosya, encoding="utf-8-sig") as f:
            for row in csv.DictReader(f):
                k = (row.get(anahtar) or "").strip().upper()
                if k:
                    out[k] = row
    except FileNotFoundError:
        pass
    return out


def sayi(x):
    try:
        s = str(x).strip().replace(",", ".")
        return float(s) if s else None
    except Exception:
        return None


def load_manuel_kaldirac():
    d = csv_oku("manuel-kaldirac.csv", "varlik")
    if not d:
        print("  ! manuel-kaldirac.csv yok -> Binance kaldirac bos kalacak.")
    return {k: sayi(v.get("max_kaldirac")) for k, v in d.items()
            if sayi(v.get("max_kaldirac")) is not None}


def kaldirac_csv_tamamla(semboller):
    mevcut = set(csv_oku("manuel-kaldirac.csv", "varlik").keys())
    eksik = [s for s in semboller if s not in mevcut]
    if not eksik:
        return
    yeni = not os.path.exists("manuel-kaldirac.csv")
    with open("manuel-kaldirac.csv", "a", encoding="utf-8") as f:
        if yeni:
            f.write("varlik,max_kaldirac\n")
        for s in eksik:
            f.write(f"{s},\n")
    print(f"  manuel-kaldirac.csv'ye {len(eksik)} varlik eklendi (bos).")


BTCTURK_ALANLAR = ["min_miktar", "min_tutar", "fiyat_adimi", "digit",
                   "miktar_adimi", "max_kaldirac", "funding_periyot"]


def load_btcturk():
    d = csv_oku("manuel-btcturk.csv", "varlik")
    if not d:
        print("  ! manuel-btcturk.csv yok -> BTCTURK satirlari bos gorunecek.")
    return d


def btcturk_csv_tamamla(semboller):
    mevcut = set(csv_oku("manuel-btcturk.csv", "varlik").keys())
    eksik = [s for s in semboller if s not in mevcut]
    if not eksik:
        return
    yeni = not os.path.exists("manuel-btcturk.csv")
    with open("manuel-btcturk.csv", "a", encoding="utf-8") as f:
        if yeni:
            f.write("varlik," + ",".join(BTCTURK_ALANLAR) + "\n")
        for s in eksik:
            f.write(s + "," * len(BTCTURK_ALANLAR) + "\n")
    print(f"  manuel-btcturk.csv'ye {len(eksik)} varlik eklendi (bos).")


def load_manuel_funding():
    d = csv_oku("manuel-funding.csv", "borsa")
    return {k.title() if k != "OKX" else "OKX": sayi(v.get("default_funding"))
            for k, v in d.items() if sayi(v.get("default_funding")) is not None}


def breakdown_text(bd):
    if not bd:
        return ""
    parts = []
    for c in bd:
        if isinstance(c, dict):
            ex, w = c.get("exchange"), c.get("weight")
        else:
            ex = c[0]
            w = c[1] if len(c) > 1 else None
        if not ex:
            continue
        parts.append(str(ex) if w is None else f"{ex}:{(w * 100 if w <= 1 else w):.0f}%")
    return ", ".join(parts)


# Kirilim yayinlamayan borsalar: hucre bos kalmasin, yontem yazilsin
YONTEM_METNI = {
    "Bybit": ("Yayinlanmiyor. Yontem: hacimce ilk 6 spot cift, 24s hacim agirlikli "
              "(agirlik = borsanin hacmi / 6 borsanin toplam hacmi), saatlik guncellenir. "
              "Havuz: Binance, OKX, Bybit, Coinbase (1. grup) + Bitget, Gate, MEXC (2. grup)."),
    "Bitget": ("Yayinlanmiyor. Yontem: en fazla 6 borsa, 24s hacim agirlikli "
               "(agirlik = borsanin hacmi / kullanilan borsalarin toplam hacmi), "
               "agirliklar 4 saatte bir guncellenir. "
               "Havuz: Bitget, Binance, Coinbase, OKX, Bybit, Gate, MEXC, Bitfinex, Kraken. "
               "Olaganustu durumda Bitget sabit agirlik atayabilir."),
}



# Index kirilimindaki kaynak borsa adlarini teklestir
KAYNAK_ESLEME = {
    "binance": "Binance", "okex": "OKX", "okx": "OKX", "coinbase": "Coinbase",
    "bybit": "Bybit", "bitget": "Bitget", "gateio": "Gate", "gate": "Gate",
    "gate.io": "Gate", "kucoin": "Kucoin", "mexc": "MEXC", "kraken": "Kraken",
    "bitfinex": "Bitfinex", "huobi": "HTX", "htx": "HTX",
    "hyperliquid": "Hyperliquid", "mexc_future": "MEXC", "bitstamp": "Bitstamp",
    "upbit": "Upbit", "crypto.com": "Crypto.com", "cryptocom": "Crypto.com",
}


def kaynak_adi(x):
    if not x:
        return None
    t = str(x).strip().lower().replace("_", "").replace(" ", "")
    return KAYNAK_ESLEME.get(t, str(x).strip().title())


def kirilim_agirliklari(bd):
    """index_breakdown -> {kaynak_borsa: yuzde}. Agirlik yoksa {kaynak: None}."""
    out = {}
    for c in (bd or []):
        if isinstance(c, dict):
            ex, w = c.get("exchange"), c.get("weight")
        else:
            ex = c[0]
            w = c[1] if len(c) > 1 else None
        ad = kaynak_adi(ex)
        if not ad:
            continue
        out[ad] = None if w is None else (w * 100 if w <= 1 else w)
    return out


def kontrat_degerler(k, sym, manuel_kaldirac, borsa):
    """Bir borsanin kontrat satirindaki 12 deger."""
    kaldirac = k.get("max_leverage")
    if kaldirac is None and borsa == "Binance":
        kaldirac = manuel_kaldirac.get(sym)
    fund = k.get("funding")
    return [
        k.get("index_price"), k.get("min_qty"), k.get("min_notional"),
        k.get("tick_size"), k.get("digit"), k.get("step_size"), kaldirac,
        (fund * 100) if fund is not None else None,
        k.get("funding_interval_h"), k.get("depth_1pct_usd"),
        k.get("spread_pct"),          # collect.py bunu ZATEN yuzde verir
        breakdown_text(k.get("index_breakdown")) or YONTEM_METNI.get(borsa, ""),
    ]


def btcturk_degerler(satir):
    """BTCTURK satiri: elle girilen alanlar 12'lik duzene oturtulur."""
    if not satir:
        return [None] * 12
    return [
        None,                              # Index Price - yok
        sayi(satir.get("min_miktar")),
        sayi(satir.get("min_tutar")),
        sayi(satir.get("fiyat_adimi")),
        sayi(satir.get("digit")),
        sayi(satir.get("miktar_adimi")),
        sayi(satir.get("max_kaldirac")),
        None,                              # Funding % - yok
        sayi(satir.get("funding_periyot")),
        None, None, "",                    # derinlik, spread, kirilim - yok
    ]


def kontrat_bicim(ws, r, c0):
    """12'lik blogun sayi bicimleri (c0 = ilk sutun)."""
    ws.cell(r, c0 + 0).number_format = '#,##0.########'
    for off in (1, 3, 5):
        ws.cell(r, c0 + off).number_format = '0.########'
    ws.cell(r, c0 + 2).number_format = '#,##0.##'
    ws.cell(r, c0 + 4).number_format = '0'
    ws.cell(r, c0 + 6).number_format = '0"x"'
    ws.cell(r, c0 + 7).number_format = '0.0000"%"'
    ws.cell(r, c0 + 8).number_format = '0.0'
    ws.cell(r, c0 + 9).number_format = '#,##0'
    ws.cell(r, c0 + 10).number_format = '0.0000"%"'


# ---------------- hacim sheet'leri (genis) ----------------

def sheet_hacim(wb, borsa, gunler, cmc, btcturk_spot=None, ilk=False):
    """EN GUNCEL gun EN SOLDA.
    Guncel gun: Fiyat | Market Cap | Open Interest | Perp Hacim (4 sutun)
    Gecmis gunler: sadece Perp Hacim (1 sutun)
    Tum veriler arsivde saklanmaya devam eder; burada sadece gosterim sadelestirilir."""
    ws = wb.active if ilk else wb.create_sheet()
    ws.title = borsa
    ws.sheet_properties.tabColor = BORSA_RENK.get(borsa, "003AFF")

    # En yeni gun basta
    gun_listesi = sorted([g for g in gunler if gunler[g].get(borsa)], reverse=True)
    veri, geriye = {}, {}
    for g in gun_listesi:
        satirlar = gunler[g].get(borsa) or []
        geriye[g] = any(s.get("geriye_donuk") for s in satirlar)
        for s in satirlar:
            veri.setdefault(s["symbol"], {})[g] = s

    # Gosterilecek varliklar = EN GUNCEL gunun ilk GOSTER_N'i.
    # Arsivde 300 varlik saklaniyor; boylece bu 150'nin gecmis gunlerdeki
    # hacimleri de dolu geliyor (o gun 150'nin disinda kalmis olsalar bile).
    GOSTER_N = 150
    if gun_listesi:
        guncel = gun_listesi[0]
        aday = [(sym, (veri[sym].get(guncel) or {}).get("perp_volume_usd") or 0)
                for sym in veri]
        aday = [a for a in aday if a[1] > 0]
        aday.sort(key=lambda x: -x[1])
        varliklar = [a[0] for a in aday[:GOSTER_N]]
        # Guncel gunde hic verisi olmayan ama gecmiste olanlari en alta ekle
        kalan = [s for s in veri if s not in set(varliklar)]
        def son_hacim(sym):
            for g in gun_listesi:
                v = (veri[sym].get(g) or {}).get("perp_volume_usd") or 0
                if v:
                    return -v
            return 0
        varliklar += sorted(kalan, key=son_hacim)[:0]   # simdilik eklenmiyor
    else:
        varliklar = []

    # Baslik: A=Varlik, B=BtcTurk Spot, sonra tarih bloklari
    for sc_, ad_ in ((1, "Varlik"), (2, "BtcTurk Spot")):
        ws.merge_cells(start_row=1, start_column=sc_, end_row=2, end_column=sc_)
        c = ws.cell(1, sc_, ad_)
        c.fill = HEAD_FILL; c.font = HEAD_FONT; c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    blok = []            # (gun, ilk_sutun, metrik_listesi)
    col = 3
    for i, g in enumerate(gun_listesi):
        metrik = GUNCEL_METRIK if i == 0 else ESKI_METRIK
        blok.append((g, col, metrik))
        if len(metrik) > 1:
            ws.merge_cells(start_row=1, start_column=col,
                           end_row=1, end_column=col + len(metrik) - 1)
        etiket = gun_bicim(g)      # geriye donuk gunlere ek isaret konmuyor
        hc = ws.cell(1, col, etiket)
        hc.fill = HEAD_FILL; hc.font = HEAD_FONT; hc.border = BORDER
        hc.alignment = Alignment(horizontal="center", vertical="center")
        for j, m in enumerate(metrik):
            sc = ws.cell(2, col + j, m)
            sc.fill = ALT_FILL; sc.font = ALT_FONT; sc.border = BORDER
            sc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        col += len(metrik)
    son_sutun = col - 1
    ws.freeze_panes = "C3"

    dolgu = PatternFill("solid", fgColor=BORSA_DOLGU.get(borsa, "FFFFFF"))
    ES_FILL = PatternFill("solid", fgColor="E8F6EE")
    HAYIR_FILL = PatternFill("solid", fgColor="FBEFEF")
    r = 3
    for sym in varliklar:
        vc = ws.cell(r, 1, sym); vc.font = FONT_B; vc.border = BORDER
        # BtcTurk spot'ta listeli mi? (temel token uzerinden bakilir: 1000PEPE -> PEPE)
        if btcturk_spot is None:
            bt = ""
        else:
            temel = temel_sembol(sym)
            bt = "Y" if (sym in btcturk_spot or (temel and temel in btcturk_spot)) else "N"
        bc = ws.cell(r, 2, bt)
        bc.font = FONT_B
        bc.border = BORDER
        bc.alignment = Alignment(horizontal="center")
        if bt == "Y": bc.fill = ES_FILL
        elif bt == "N": bc.fill = HAYIR_FILL
        for g, c0, metrik in blok:
            s = veri[sym].get(g) or {}
            if len(metrik) > 1:
                fiyat, mcap = cmc_bak(cmc, g, sym) if s else (None, None)
                vals = [fiyat, mcap, s.get("open_interest_usd"), s.get("perp_volume_usd")]
            else:
                vals = [s.get("perp_volume_usd")]
            for j, v in enumerate(vals):
                cell = ws.cell(r, c0 + j, v)
                cell.font = FONT; cell.border = BORDER; cell.fill = dolgu
            if len(metrik) > 1:
                ws.cell(r, c0).number_format = '#,##0.########'
                for j in (1, 2, 3):
                    ws.cell(r, c0 + j).number_format = '#,##0'
            else:
                ws.cell(r, c0).number_format = '#,##0'
        r += 1

    ws.column_dimensions["A"].width = 13
    ws.column_dimensions["B"].width = 13
    for g, c0, metrik in blok:
        genislik = [14, 18, 18, 18] if len(metrik) > 1 else [18]
        for j, w in enumerate(genislik):
            ws.column_dimensions[get_column_letter(c0 + j)].width = w
    if r > 3:
        ws.auto_filter.ref = f"A2:{get_column_letter(son_sutun)}{r - 1}"
    return len(varliklar), len(gun_listesi)


def sheet_total(wb, gunler, cmc):
    """A: Varlik, B+: tarihler. Deger = 6 borsanin TOPLAM perp hacmi."""
    ws = wb.create_sheet(title="Total")
    ws.sheet_properties.tabColor = "0A9D57"
    gun_listesi = sorted(gunler, reverse=True)   # en guncel gun EN SOLDA
    toplam = {}   # sym -> {gun: toplam}
    for g in gun_listesi:
        for borsa in BORSA_SIRASI:
            for s in (gunler[g].get(borsa) or []):
                v = s.get("perp_volume_usd") or 0
                if v:
                    d = toplam.setdefault(s["symbol"], {})
                    d[g] = d.get(g, 0) + v

    def anahtar(sym):
        for g in gun_listesi:                     # en guncel gunden geriye
            if toplam[sym].get(g):
                return -toplam[sym][g]
        return 0
    varliklar = sorted(toplam.keys(), key=anahtar)

    basliklar = ["Varlik"] + [gun_bicim(g) for g in gun_listesi]
    for c, h in enumerate(basliklar, 1):
        cell = ws.cell(1, c, h)
        cell.fill = HEAD_FILL; cell.font = HEAD_FONT; cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "B2"

    r = 2
    for sym in varliklar:
        vc = ws.cell(r, 1, sym); vc.font = FONT_B; vc.border = BORDER
        for i, g in enumerate(gun_listesi):
            cell = ws.cell(r, 2 + i, toplam[sym].get(g))
            cell.font = FONT; cell.border = BORDER
            cell.number_format = '#,##0'
        r += 1
    ws.column_dimensions["A"].width = 13
    for i in range(len(gun_listesi)):
        ws.column_dimensions[get_column_letter(2 + i)].width = 20
    if r > 2:
        ws.auto_filter.ref = f"A1:{get_column_letter(1 + len(gun_listesi))}{r - 1}"
    return len(varliklar)


# ---------------- kontrat sheet'leri ----------------


# Varlik sheet'lerinin altina konan alan aciklamalari
ALAN_ACIKLAMA = [
    ("Index Price",
     "Borsanin dis spot piyasalardan derledigi referans fiyat. Funding ve likidasyon buna gore isler. "
     "Borsanin kendi order book fiyati degildir."),
    ("Min Miktar",
     "Acabilecegin en kucuk pozisyon miktari (coin cinsinden). Ornek: 0,001 ise en az 0,001 BTC."),
    ("Min Tutar ($)",
     "Pozisyonun en az kac dolar degerinde olmasi gerektigi. Min Miktar ile birlikte gecerlidir; "
     "hangisi daha kisitlayiciysa o baglar."),
    ("Fiyat Adimi",
     "Emir fiyatinin hangi araliklarla girilebilecegi (tick size). 0,1 ise 66420,1 girilir ama "
     "66420,15 girilemez. Emir defterinin cozunurlugunu belirler."),
    ("Digit",
     "Fiyat adiminin kac ondalik haneye denk geldigi. Adim 0,1 ise 1 digit; 0,01 ise 2 digit."),
    ("Miktar Adimi",
     "Pozisyon miktarinin hangi araliklarla girilebilecegi (step size). 0,001 ise 0,0015 girilemez. "
     "Min Miktar alt siniri, bu ise basamak yuksekligini belirler."),
    ("Max Kaldirac",
     "O varlikta izin verilen en yuksek kaldirac. Dusuk kaldirac genelde borsanin o varligi "
     "riskli gordugunu gosterir."),
    ("Funding %",
     "En son funding orani. Pozitifse long pozisyonlar short'lara oder, negatifse tersi. "
     "Talep yonunun gostergesidir."),
    ("Funding Periyot (saat)",
     "Funding'in kac saatte bir odendigi. Cogu borsada 8 saat, Hyperliquid'de 1 saat."),
    ("Derinlik +-1% ($)",
     "Fiyatin %1 alti ve ustundeki emirlerin toplam dolar degeri. Likidite gostergesi: "
     "yuksekse buyuk emir fiyati az kaydirir, dusukse piyasa sigdir."),
    ("Spread %",
     "En iyi alis ile en iyi satis arasindaki fark, yuzde olarak. Dar spread likit piyasa demektir."),
    ("Index Kirilimi",
     "Index fiyatinin hangi borsalardan hangi agirlikla hesaplandigi. "
     "Bybit ve Bitget bu bilgiyi yayinlamadigi icin bos kalabilir."),
]


def sheet_kontrat_varlik(wb, kayit, tarihler, manuel_kaldirac, btcturk):
    """Bir varlik icin: satirlar borsalar (+BTCTURK), sutunlar 12 kontrat alani."""
    sym = kayit["symbol"]
    ws = wb.create_sheet(title=sym[:31])
    basliklar = ["Borsa", "Borsa Sembolu", "Veri Tarihi"] + KONTRAT_SUTUNLAR
    for c, h in enumerate(basliklar, 1):
        cell = ws.cell(1, c, h)
        cell.fill = HEAD_FILL; cell.font = HEAD_FONT; cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"

    r = 2
    for borsa in BORSA_SIRASI:
        k = (kayit.get("exchanges") or {}).get(borsa) or {}
        bc = ws.cell(r, 1, borsa); bc.font = FONT_B
        ws.cell(r, 2, k.get("borsa_sembolu") or "").font = FONT
        ws.cell(r, 3, tarihler.get(borsa) or "").font = FONT
        vals = kontrat_degerler(k, sym, manuel_kaldirac, borsa)
        for j, v in enumerate(vals):
            cell = ws.cell(r, 4 + j, v)
            cell.font = FONT
        dolgu = PatternFill("solid", fgColor=BORSA_DOLGU.get(borsa, "FFFFFF"))
        for c in range(1, len(basliklar) + 1):
            ws.cell(r, c).border = BORDER
            ws.cell(r, c).fill = dolgu
        kontrat_bicim(ws, r, 4)
        r += 1

    # Kendi borsan
    bc = ws.cell(r, 1, KENDI_BORSA); bc.font = FONT_B
    ws.cell(r, 2, "").font = FONT
    ws.cell(r, 3, "elle girilen").font = FONT
    for j, v in enumerate(btcturk_degerler(btcturk.get(sym))):
        ws.cell(r, 4 + j, v).font = FONT
    dolgu = PatternFill("solid", fgColor=BORSA_DOLGU[KENDI_BORSA])
    for c in range(1, len(basliklar) + 1):
        ws.cell(r, c).border = Border(left=THIN, right=THIN,
                                     top=Side(style="medium", color="003AFF"),
                                     bottom=Side(style="medium", color="003AFF"))
        ws.cell(r, c).fill = dolgu
    kontrat_bicim(ws, r, 3)

    # ---- INDEX KIRILIMI MATRISI (formulde kullanilabilir) ----
    kayitlar = {}
    kaynaklar = []
    for borsa in BORSA_SIRASI:
        k = (kayit.get("exchanges") or {}).get(borsa) or {}
        ag = kirilim_agirliklari(k.get("index_breakdown"))
        kayitlar[borsa] = ag
        for ad in ag:
            if ad not in kaynaklar:
                kaynaklar.append(ad)
    mr = r + 2
    if kaynaklar:
        bas = ws.cell(mr, 1, "Index kirilimi - agirlik matrisi (%)")
        bas.font = Font(name="Arial", size=11, bold=True, color="003AFF")
        mr += 1
        ws.cell(mr, 1, "Index sahibi").font = FONT_B
        ws.cell(mr, 1).fill = ALT_FILL; ws.cell(mr, 1).border = BORDER
        for j, ad in enumerate(kaynaklar):
            hc = ws.cell(mr, 2 + j, ad)
            hc.font = ALT_FONT; hc.fill = ALT_FILL; hc.border = BORDER
            hc.alignment = Alignment(horizontal="center", wrap_text=True)
        tc = ws.cell(mr, 2 + len(kaynaklar), "TOPLAM")
        tc.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
        tc.fill = HEAD_FILL; tc.border = BORDER
        tc.alignment = Alignment(horizontal="center")
        mr += 1
        for borsa in BORSA_SIRASI:
            ag = kayitlar.get(borsa) or {}
            ws.cell(mr, 1, borsa).font = FONT_B
            ws.cell(mr, 1).border = BORDER
            ws.cell(mr, 1).fill = PatternFill("solid", fgColor=BORSA_DOLGU.get(borsa, "FFFFFF"))
            for j, ad in enumerate(kaynaklar):
                if ad not in ag:
                    # o borsa bu kaynagi HIC kullanmiyor -> bos
                    c_ = ws.cell(mr, 2 + j)
                    c_.border = BORDER
                    continue
                deger = ag.get(ad)
                if deger is None:
                    # kaynak olarak kullaniyor AMA agirligi yayinlamiyor
                    c_ = ws.cell(mr, 2 + j, "K")
                    c_.font = Font(name="Arial", size=9, bold=True, color="35496B")
                    c_.alignment = Alignment(horizontal="center")
                    c_.fill = PatternFill("solid", fgColor="EEF3FB")
                else:
                    c_ = ws.cell(mr, 2 + j, deger)
                    c_.font = FONT
                    c_.number_format = '0.00'
                c_.border = BORDER
            sayisal = [v for v in ag.values() if isinstance(v, (int, float))]
            if sayisal:
                tt = ws.cell(mr, 2 + len(kaynaklar), sum(sayisal))
                tt.number_format = '0.00'
                tt.fill = PatternFill("solid",
                                      fgColor="FBEFEF" if abs(sum(sayisal) - 100) > 1 else "E8F6EE")
            elif ag:
                tt = ws.cell(mr, 2 + len(kaynaklar), f"{len(ag)} kaynak")
                tt.fill = PatternFill("solid", fgColor="EEF3FB")
            else:
                tt = ws.cell(mr, 2 + len(kaynaklar))
            tt.font = FONT_B; tt.border = BORDER
            mr += 1
        ws.cell(mr, 1, "Sayi = yayinlanan agirlik %.   K = o borsa bu kaynagi kullaniyor "
                       "ama agirligini yayinlamiyor (Gate, MEXC).   "
                       "Bos = o kaynak hic kullanilmiyor.   "
                       "Bybit ve Bitget kaynak listesini de yayinlamaz.").font = \
            Font(name="Arial", size=9, italic=True, color="5B6B85")
        mr += 1

    # ---- alan aciklamalari ----
    ar = mr + 2
    bas = ws.cell(ar, 1, "Sutunlar ne anlama geliyor?")
    bas.font = Font(name="Arial", size=11, bold=True, color="003AFF")
    ar += 1
    for ad, metin in ALAN_ACIKLAMA:
        ws.cell(ar, 1, ad).font = FONT_B
        c = ws.cell(ar, 3, metin)
        c.font = FONT
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=ar, start_column=3, end_row=ar, end_column=10)
        ws.row_dimensions[ar].height = 28
        ar += 1

    for c, w in zip("AB", [16, 12]):
        ws.column_dimensions[c].width = w
    for j, w in enumerate([13, 11, 11, 11, 6, 11, 9, 10, 9, 15, 9, 34]):
        ws.column_dimensions[get_column_letter(3 + j)].width = w
    return r


def sheet_analiz(wb, varliklar, tarihler, manuel_kaldirac, btcturk):
    """Tek tablo: Varlik | Borsa | 12 alan. Filtreden varlik secilir."""
    ws = wb.create_sheet(title="Analiz")
    ws.sheet_properties.tabColor = "E6A100"
    basliklar = ["Varlik", "Borsa", "Borsa Sembolu", "Veri Tarihi"] + KONTRAT_SUTUNLAR
    for c, h in enumerate(basliklar, 1):
        cell = ws.cell(1, c, h)
        cell.fill = HEAD_FILL; cell.font = HEAD_FONT; cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "C2"

    r = 2
    for a in varliklar:
        sym = a["symbol"]
        for borsa in BORSA_SIRASI + [KENDI_BORSA]:
            ws.cell(r, 1, sym).font = FONT_B
            ws.cell(r, 2, borsa).font = FONT_B
            if borsa == KENDI_BORSA:
                ws.cell(r, 3, "").font = FONT
                ws.cell(r, 4, "elle girilen").font = FONT
                vals = btcturk_degerler(btcturk.get(sym))
            else:
                k = (a.get("exchanges") or {}).get(borsa) or {}
                ws.cell(r, 3, k.get("borsa_sembolu") or "").font = FONT
                ws.cell(r, 4, tarihler.get(borsa) or "").font = FONT
                vals = kontrat_degerler(k, sym, manuel_kaldirac, borsa)
            for j, v in enumerate(vals):
                ws.cell(r, 5 + j, v).font = FONT
            dolgu = PatternFill("solid", fgColor=BORSA_DOLGU.get(borsa, "FFFFFF"))
            for c in range(1, len(basliklar) + 1):
                ws.cell(r, c).border = BORDER
                ws.cell(r, c).fill = dolgu
            kontrat_bicim(ws, r, 5)
            r += 1

    for c, w in zip("ABCD", [13, 13, 14, 12]):
        ws.column_dimensions[c].width = w
    for j, w in enumerate([13, 11, 11, 11, 6, 11, 9, 10, 9, 15, 9, 34]):
        ws.column_dimensions[get_column_letter(5 + j)].width = w
    if r > 2:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(basliklar))}{r - 1}"
    return r - 2


# ---------------- notlar ----------------


def sheet_notlar(wb, hacim_gun, kontrat_tarihler, manuel_kaldirac, manuel_funding, btcturk):
    ws = wb.create_sheet(title="Notlar")
    ws.cell(1, 1, "MarketPano - notlar ve kaynaklar").font = Font(name="Arial", size=12, bold=True)
    satirlar = [
        ("Yapı", ""),
        ("Borsa sheet'leri (6 adet)",
         "A sütunu Varlık. Sonra her gün için 4 sütunluk blok: Fiyat, Market Cap, "
         "Perp Hacim, Open Interest. Yeni gün sağa eklenir."),
        ("Satır sıralaması", "En son günün perp hacmine göre azalan"),
        ("Kapsam", "O borsanın kendi 24 saatlik perp hacminde ilk 150 kripto varlığı"),
        ("Total sheet", "A sütunu Varlık, sonraki sütunlar tarih. Değer: 6 borsanın TOPLAM perp hacmi."),
        ("Varlık sheet'leri (50 adet)",
         "CMC ilk 50. Her sheet'te borsalar alt alta, sütunlar 12 kontrat alanı, en altta BTCTURK."),
        ("Analiz sheet'i",
         "Varlık | Borsa | 12 alan. Filtreden varlık seçince o varlığın tüm borsaları alt alta gelir."),
        ("Arşivdeki gün sayısı", f"{hacim_gun} gün"),
        ("", ""),
        ("Veri kaynakları", ""),
        ("Perp hacim ve Open Interest", "İlgili borsanın public API'si"),
        ("Fiyat ve Market Cap", "CoinMarketCap ilk 1000"),
        ("Kontrat alanları", "İlgili borsanın public API'si"),
        ("Derinlik +-1% ve Spread", "Order book'tan hesaplanır (ham dolar, etiket yok)"),
        ("Tarih başlığında (mum) yazıyorsa",
         "O gün o borsanın verisi geriye dönük dolduruldu: UTC günlük mum. "
         "Normal kayıtlar o andaki son 24 saattir. Aynı borsa, biraz farklı pencere."),
        ("", ""),
        ("Süzgeçler", ""),
        ("Hisse/emtia süzgeci",
         "Borsalar tokenize hisse ve emtia perp'leri de listeliyor (SNDK, SOXL, XAU, MU...). "
         "CMC'de olmadıkları için listeden çıkarılır; sadece kripto kalır."),
        ("Çarpanlı semboller",
         "Binance/Bybit 1000PEPE, 1000SHIB gibi; Hyperliquid kPEPE gibi listeler "
         "(1 kontrat = 1000 token). Kripto sayılır, elenmez. Fiyat/Market Cap temel "
         "token'dan alınır, böylece borsalar arası karşılaştırılabilir."),
        ("", ""),
        ("Kontrat verisi", ""),
        ("Güncelleme", "TEK FOTOĞRAF - her gün yenilenmez. 'MarketPano kontrat guncelle' akışı elle çalıştırılır."),
        ("Çekildiği tarih",
         " · ".join(f"{b}: {g}" for b, g in sorted(kontrat_tarihler.items())) or "veri yok"),
        ("Index kırılımı - Binance, OKX, Gate", "API'den (kaynak borsa + ağırlık)"),
        ("Index kırılımı - Bitget", "Bitget canlı ağırlıkları yayınlamıyor. Yöntemi aşağıda."),
        ("Index kırılımı - Bybit", "Bybit canlı ağırlıkları yayınlamıyor. Yöntemi aşağıda."),
        ("", ""),
        ("BYBIT index hesaplama yöntemi", ""),
        ("Formül",
         "Index = Σ (spot fiyat × ağırlık). Ağırlık = o borsanın 24s hacmi ÷ altı borsanın toplam 24s hacmi."),
        ("Kaç kaynak", "Hacimce ilk 6 spot çift"),
        ("Ağırlık güncelleme", "Saatlik"),
        ("Kaynak havuzu",
         "1. grup: Binance, OKX, Bybit, Coinbase — 2. grup: Bitget, Gate, MEXC"),
        ("Fiyat kaynağı",
         "Normalde son işlem fiyatı. İşlem az ya da fiyat anormalse emir defterinden hesaplanır: "
         "(Ask1 × BidHacim1 + Bid1 × AskHacim1) ÷ (BidHacim1 + AskHacim1)"),
        ("Koruma",
         "Medyandan %5 sapan bileşen dışlanır, ağırlığı yumuşatma algoritmasıyla azaltılıp "
         "diğerlerine dağıtılır. BTC ve ETH'te eşik %1, XAU/XAG'de %3. 15 dakika işlem görmeyen çift dışlanır."),
        ("Geri dönüş şartı",
         "Binance/OKX/Bybit/Coinbase'ten en az biri, ya da Bitget/Gate/MEXC'ten en az ikisi bulunmalı "
         "ve hacim ağırlıklı toplam >= %55 olmalı."),
        ("", ""),
        ("BITGET index hesaplama yöntemi", ""),
        ("Formül",
         "Index = Σ (spot fiyat × ağırlık). Ağırlık = o borsanın 24s hacmi ÷ kullanılan borsaların toplam 24s hacmi."),
        ("Kaç kaynak", "En fazla 6 borsa"),
        ("Ağırlık güncelleme", "4 saatte bir. Index fiyatının kendisi en az 200 ms'de bir güncellenir."),
        ("Kaynak havuzu",
         "Bitget, Binance, Coinbase, OKX, Bybit, Gate, MEXC, Bitfinex, Kraken. "
         "NOT: Bu liste Bitget destek makalesinin bölgesel bir sürümünden alındı; "
         "ana sayfada yalnızca 'başlıca borsalar' deniyor, isim listesi verilmiyor."),
        ("Koruma - sapma",
         "Medyandan %5 sapan bileşen dışlanır; medyanın %2'sine dönene kadar geri alınmaz."),
        ("Koruma - durgunluk",
         "15 dakika fiyat güncellemeyen borsa otomatik çıkarılır; medyanın %2'si içinde "
         "güncellemeye başlayınca geri alınır."),
        ("Olağanüstü durum",
         "Bitget sistemik riski önlemek için bir borsayı tamamen çıkarabilir YA DA SABİT AĞIRLIK "
         "atayabilir. Normal işleyişte ağırlık tabanı yoktur, ancak bu yetki saklıdır."),
        ("Kaynak değişimi",
         "Kaynak listesi değişimi index'i %0,1'den fazla oynatacaksa geçiş kademeli yapılır."),
        ("Çapraz kur",
         "Bir borsada istenen kotasyon yoksa çevrim yapılır (örn. Coinbase BTC/USD → BTC/USDC)."),
        ("Dış kaynak yoksa",
         "Vadeli emir defterinden derinlik ağırlıklı orta fiyat türetilir; "
         "Index(Tn) = a x orta fiyat + (1-a) x Index(Tn-1), a varsayilan 0,1818."),
        ("Doğrulama",
         "Bitget metodolojisi 6 Ağustos 2026'da resmi destek sayfasından doğrudan okundu "
         "(sayfa tarihi 13 Mart 2025). Bybit metodolojisi resmi yardım merkezinden okundu "
         "(sayfa 19 Haziran 2026'da güncellenmiş)."),
        ("Index kırılımı - Hyperliquid", "Sabit formül, aşağıda"),
        ("Hyperliquid yöntemi", "AĞIRLIKLI MEDYAN (ortalama DEĞİL) - ağırlıklar medyan oyudur"),
        ("Hyperliquid - normal varlıklar (örn. BTC)",
         "Binance 3 (%27,3) · OKX 2 (%18,2) · Bybit 2 (%18,2) · Kraken 1 (%9,1) · "
         "Kucoin 1 (%9,1) · Gate 1 (%9,1) · MEXC 1 (%9,1) — Hyperliquid hariç"),
        ("Hyperliquid - ana likiditesi kendisinde olanlar (örn. HYPE)",
         "Sadece Hyperliquid; dış kaynaklar yeterli likiditeye kadar dahil edilmez"),
        ("", ""),
        ("Index koruma kuralları (borsa dokümanlarından)", ""),
        ("Binance",
         "Bir kaynağın fiyatı medyandan %5'ten fazla saparsa, o fiyat medyanın 1,05 / 0,95 katına "
         "KISILIR (dışlanmaz). 10 saniye veri gelmezse o kaynağın ağırlığı sıfırlanır."),
        ("OKX",
         "3'ten fazla geçerli kaynak varsa EŞİT ağırlıklı ortalama; medyandan %3'ten fazla sapan "
         "kaynak medyanın 0,97-1,03 bandına kısılır (bazı endekslerde eşik %2). "
         "2 kaynak kalırsa eşit ağırlık, 1 kaynak kalırsa o fiyat doğrudan kullanılır."),
        ("Bybit",
         "Medyandan %5'ten fazla sapan bileşen geçici olarak DIŞLANIR; ağırlığı yumuşatma "
         "algoritmasıyla azaltılıp diğerlerine dağıtılır. BTC ve ETH'te eşik %1, XAU/XAG'de %3. "
         "15 dakika işlem görmeyen çift dışlanır. Yeniden dahil olmak için hacim ağırlıklı "
         "toplamın >=%55 olması gibi şartlar aranır."),
        ("Gate",
         "Medyandan %8'ten fazla sapan bileşen dışlanır. %2-%8 arası sapmada fiyat medyanın "
         "%98 / %102'sine kısılır. 2 kaynak kalırsa anormal olan elenir, 1 kaynak kalırsa "
         "o fiyat doğrudan kullanılır."),
        ("Bitget", "Public dokümanda net kural bulunamadı."),
        ("MEXC",
         "Formül: Index = Σ(Ağırlık% × Spot Fiyat). Ağırlık% = o borsanın ağırlığı ÷ toplam ağırlık. "
         "Bileşenler ve ağırlıklar periyodik güncellenir; MEXC ağırlık DEĞERLERİNİ yayınlamaz, "
         "yalnızca kaynak borsa listesini API'de verir (indexOrigin). "
         "Koruma: medyandan %1 sapan kaynak dışlanır (üç kaynaktan az kalırsa uygulanmaz); "
         "verisi geciken ya da sapan borsa çıkarılır, düzelince geri alınır. "
         "Bu %1 eşiği yedi borsa içindeki en sıkı eşiktir."),
        ("Hyperliquid", "Ağırlıklı medyan kullanır; sapma kısıtlaması yerine medyanın kendisi koruma sağlar."),
        ("ÖNEMLİ AYRIM",
         "Bu kuralların hepsi FİYAT SAPMASI ile ilgilidir. Hiçbir borsada 'bir kaynağın ağırlığı "
         "şu yüzdenin altına düşemez' şeklinde bir AĞIRLIK TABANI kuralı bulunmadı. Ağırlıklar "
         "hacme göre serbestçe değişir; kurallar sadece sapan fiyatı kısar ya da kaynağı dışlar."),
        ("", ""),
        ("", ""),
        ("Default funding (faiz bileşeni, borsa başına sabit)", ""),
    ]
    for b in BORSA_SIRASI:
        d = manuel_funding.get(b)
        satirlar.append(("   " + b, f"%{d} / 8 saat" if d is not None else "teyit bekliyor / girilmedi"))
    satirlar += [
        ("", ""),
        ("Elle girilen veriler", ""),
        ("manuel-kaldirac.csv", f"Binance max kaldıraç (API'de yok) - {len(manuel_kaldirac)} varlık dolu"),
        ("manuel-btcturk.csv", f"BTCTURK kendi kontrat değerleri - {len(btcturk)} varlık satırı"),
        ("manuel-funding.csv", "Default funding değerleri (borsa başına)"),
        ("Önemli", "Elle girilen değerler tarih bazlı değildir; her yerde aynı görünür"),
        ("", ""),
        ("Excel nasıl üretilir", ""),
        ("Yöntem", "Her çalıştırmada arşivden sıfırdan üretilir; dosyaya elle yazılanlar korunmaz"),
        ("Elle veri girmek için", "manuel-*.csv dosyalarını GitHub'da düzenle"),
    ]
    r = 3
    for a, b in satirlar:
        ca, cb = ws.cell(r, 1, a), ws.cell(r, 2, b)
        ca.font = FONT_B if (a and not b) else FONT
        cb.font = FONT
        cb.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    ws.column_dimensions["A"].width = 56
    ws.column_dimensions["B"].width = 95
    return ws


# ---------------- ana akis ----------------

def main():
    gunler = hacim_gunleri()
    cmc = cmc_gunleri()
    btcturk_spot = btcturk_spot_seti()
    kontrat_varliklar, kontrat_tarihler = kontrat_snapshot()

    if not gunler and not kontrat_varliklar:
        print("! Ne hacim ne kontrat verisi bulundu. Once veri cekilmeli.")
        return 1

    semboller = [a["symbol"] for a in kontrat_varliklar]
    if semboller:
        kaldirac_csv_tamamla(semboller)
        btcturk_csv_tamamla(semboller)
    manuel_kaldirac = load_manuel_kaldirac()
    manuel_funding = load_manuel_funding()
    btcturk = load_btcturk()

    wb = openpyxl.Workbook()
    ozet = []
    for i, borsa in enumerate(BORSA_SIRASI):
        adet, gun_adet = sheet_hacim(wb, borsa, gunler, cmc, btcturk_spot, ilk=(i == 0))
        ozet.append(f"{borsa}:{adet}v/{gun_adet}g")
    total_adet = sheet_total(wb, gunler, cmc)
    for a in kontrat_varliklar:
        sheet_kontrat_varlik(wb, a, kontrat_tarihler, manuel_kaldirac, btcturk)
    analiz_satir = sheet_analiz(wb, kontrat_varliklar, kontrat_tarihler,
                                manuel_kaldirac, btcturk)
    sheet_notlar(wb, len(gunler), kontrat_tarihler, manuel_kaldirac,
                 manuel_funding, btcturk)

    wb.save(OUT)
    print(f"[EXCEL] {OUT} yazildi.")
    print(f"  Sheet sayisi   : {len(wb.sheetnames)}")
    print(f"  Borsa sheet'i  : {' | '.join(ozet)}")
    print(f"  Total          : {total_adet} varlik")
    print(f"  Varlik sheet'i : {len(kontrat_varliklar)}")
    print(f"  Analiz satir   : {analiz_satir}")
    print(f"  Hacim arsivi   : {len(gunler)} gun | cmc {len(cmc)} gun")
    print(f"  Kontrat tarihi : {kontrat_tarihler or '-'}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
